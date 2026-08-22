import csv
import io
import re
from pathlib import Path

from openpyxl import load_workbook

RESOURCE_MARKER = 'SD'

DATE_HEADING_PATTERN = re.compile(
    r'\bdate\b|\bcreated\s*on\b|\bweek\s*ending\b|\bperiod\b',
    re.IGNORECASE,
)

# Known column headings → field type (before the SD resource block).
FIELD_TYPE_OVERRIDES = {
    'ticket id': 'number',
    'created on': 'date',
    'description': 'text',
    'team': 'text',
    'status': 'text',
    'business user': 'text',
    'assigned': 'text',
    'comment': 'text',
    'transport number': 'text',
    'series number': 'text',
}


def _normalize_heading(value):
    if value is None:
        return ''
    return str(value).strip()


def _detect_data_field_type(heading):
    key = heading.lower()
    if key in FIELD_TYPE_OVERRIDES:
        return FIELD_TYPE_OVERRIDES[key]
    if DATE_HEADING_PATTERN.search(heading):
        return 'date'
    if re.fullmatch(r'ticket\s*id', key):
        return 'number'
    return 'text'


def _find_resource_start_index(headings):
    for index, heading in enumerate(headings):
        if heading.upper() == RESOURCE_MARKER:
            return index
    return None


def build_columns_from_headings(headings):
    cleaned = [_normalize_heading(h) for h in headings]
    cleaned = [h for h in cleaned if h]
    if not cleaned:
        raise ValueError('Spreadsheet must contain column headings in the first row.')

    resource_start = _find_resource_start_index(cleaned)
    if resource_start is None:
        raise ValueError(
            f'Spreadsheet must include a "{RESOURCE_MARKER}" column marking '
            'where developer resource hour fields begin.'
        )

    columns = []
    for index, heading in enumerate(cleaned):
        if index >= resource_start:
            columns.append({
                'name': heading,
                'type': 'hours',
                'is_resource': True,
            })
        else:
            columns.append({
                'name': heading,
                'type': _detect_data_field_type(heading),
                'is_resource': False,
            })

    return columns, resource_start


def _read_xlsx_headings(uploaded_file):
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    workbook.close()
    if not first_row:
        raise ValueError('Spreadsheet is empty.')
    return list(first_row)


def _read_csv_headings(uploaded_file):
    raw = uploaded_file.read()
    if isinstance(raw, bytes):
        text = raw.decode('utf-8-sig')
    else:
        text = raw

    reader = csv.reader(io.StringIO(text))
    first_row = next(reader, None)
    if not first_row:
        raise ValueError('Spreadsheet is empty.')
    return first_row


def parse_spreadsheet_headings(uploaded_file):
    filename = getattr(uploaded_file, 'name', '') or ''
    extension = Path(filename).suffix.lower()

    if extension == '.csv':
        headings = _read_csv_headings(uploaded_file)
        uploaded_file.seek(0)
    elif extension in {'.xlsx', '.xlsm'}:
        headings = _read_xlsx_headings(uploaded_file)
        uploaded_file.seek(0)
    else:
        raise ValueError('Upload a .xlsx or .csv file with headings in the first row.')

    return build_columns_from_headings(headings)
